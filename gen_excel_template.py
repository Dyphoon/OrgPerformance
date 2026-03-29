#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
import os

def style_header(cell, bg='2E4057', fg='FFFFFF'):
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_title(ws, text, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=14, color='2E4057')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30

def thin_border():
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)

def apply_border(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = thin_border()

wb = Workbook()

# ============================================================
# Sheet 1: 模版页
# ============================================================
ws1 = wb.active
ws1.title = '模版页'
ws1.sheet_view.showGridLines = False

style_title(ws1, '绩效考核模版页', 1)
ws1.row_dimensions[1].height = 35

# 第2行：维度说明
ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=17)
cell = ws1.cell(row=2, column=1,
    value='说明：以下为某银行分行绩效考核指标模版。指标计算由Excel公式自动完成，系统读取计算结果落库。')
cell.font = Font(italic=True, size=9, color='888888')
cell.alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[2].height = 20

# 第3行：表头
headers = [
    '维度', '类别', '一级指标', '二级指标', '权重', '单位',
    '全年目标', '进度目标', '实际值', '全年完成率', '进度完成率',
    '指标百分制得分', '指标权重得分', '类别得分', '维度得分', '总得分'
]
for i, h in enumerate(headers, 1):
    c = ws1.cell(row=3, column=i, value=h)
    style_header(c)
ws1.row_dimensions[3].height = 35

# 示例数据（2行）
sample_data = [
    # 维度, 类别, 一级指标, 二级指标, 权重, 单位, 全年目标, 进度目标, 实际值, 全年完成率, 进度完成率, 百分制, 权重得分, 类别得分, 维度得分, 总得分
    ['业务发展', '存款业务', '存款规模', '日均存款余额', '0.15', '亿元', '100.00', '8.33', '7.85', '=I4/G4', '=I4/H4', '=MAX(0,MIN(100,I4/H4*100))', '=L4*E4', '=SUM(M4:M5)', '=SUM(N4:N5)', '=SUM(O:O)'],
    ['业务发展', '存款业务', '存款规模', '存款增长率', '0.10', '%', '15.00', '1.25', '1.10', '=I5/G5', '=I5/H5', '=MAX(0,MIN(100,I5/H5*100))', '=L5*E5', '=SUM(M4:M5)', '=SUM(N4:N5)', '=SUM(O:O)'],
    ['风险控制', '资产质量', '不良贷款率', '不良贷款率控制', '0.20', '%', '1.50', '0.125', '0.11', '=I6/G6', '=I6/H6', '=MAX(0,MIN(100,(1-I6/G6)*100))', '=L6*E6', '=SUM(M6:M7)', '=SUM(N4:N7)', '=SUM(O:O)'],
    ['风险控制', '资产质量', '拨备覆盖率', '拨备覆盖率', '0.15', '%', '180.00', '15.00', '16.50', '=I7/G7', '=I7/H7', '=MAX(0,MIN(100,I7/G7*100))', '=L7*E7', '=SUM(M6:M7)', '=SUM(N4:N7)', '=SUM(O:O)'],
]

# 填充示例数据行
for ri, row_data in enumerate(sample_data, 4):
    for ci, val in enumerate(row_data, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.border = thin_border()
        if ci == 5:  # 权重列
            c.number_format = '0.00%'
        elif ci in [7, 8, 9]:  # 数值列
            c.number_format = '0.00'
        elif ci in [10, 11, 12, 13, 14, 15, 16]:
            c.number_format = '0.0000'
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[ri].height = 22

# 列宽
col_widths_1 = [12, 12, 16, 16, 8, 8, 12, 12, 10, 12, 12, 14, 12, 10, 10, 10]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 2: 机构页
# ============================================================
ws2 = wb.create_sheet('机构页')
ws2.sheet_view.showGridLines = False

style_title(ws2, '被评价机构列表', 1)
ws2.row_dimensions[1].height = 35

headers2 = ['机构名称', '机构ID', '分组名称', '机构负责人']
for i, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=i, value=h)
    style_header(c)
ws2.row_dimensions[2].height = 30

institutions = [
    ['北京分行', 'ORG_BJ', '北方区', '张三/EMP001'],
    ['上海分行', 'ORG_SH', '华东区', '李四/EMP002'],
    ['深圳分行', 'ORG_SZ', '华南区', '王五/EMP003'],
    ['成都分行', 'ORG_CD', '西南区', '赵六/EMP004'],
    ['杭州分行', 'ORG_HZ', '华东区', '孙七/EMP005'],
]
for ri, row_data in enumerate(institutions, 3):
    for ci, val in enumerate(row_data, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[ri].height = 22

col_widths_2 = [20, 15, 15, 20]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 3: 数据收集页
# ============================================================
ws3 = wb.create_sheet('数据收集页')
ws3.sheet_view.showGridLines = False

style_title(ws3, '数据收集页（用于数据填写与分发）', 1)
ws3.row_dimensions[1].height = 35

# 表头3行
# 第2行：收数指标名称
row2 = ['机构名称', '日均存款余额', '存款增长率', '不良贷款率', '拨备覆盖率']
# 第3行：收数人
row3 = ['收数人', '张三/EMP001', '张三/EMP001', '李四/EMP002', '李四/EMP002']
# 第4行：指标单位
row4 = ['指标单位', '亿元', '%', '%', '%']

for ci, val in enumerate(row2, 1):
    c = ws3.cell(row=2, column=ci, value=val)
    if ci == 1:
        style_header(c, '2E4057')
    else:
        style_header(c, '048A81')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for ci, val in enumerate(row3, 1):
    c = ws3.cell(row=3, column=ci, value=val)
    if ci == 1:
        style_header(c, '2E4057')
    else:
        style_header(c, '048A81')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for ci, val in enumerate(row4, 1):
    c = ws3.cell(row=4, column=ci, value=val)
    if ci == 1:
        style_header(c, '2E4057')
    else:
        style_header(c, '048A81')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws3.row_dimensions[2].height = 30
ws3.row_dimensions[3].height = 30
ws3.row_dimensions[4].height = 30

# 数据行（机构 × 指标 二维表）
for ri, (org_name, org_id) in enumerate([
    ('北京分行', 'ORG_BJ'), ('上海分行', 'ORG_SH'), ('深圳分行', 'ORG_SZ'),
    ('成都分行', 'ORG_CD'), ('杭州分行', 'ORG_HZ'),
], 5):
    c1 = ws3.cell(row=ri, column=1, value=org_name)
    style_header(c1, '2E4057')
    c1.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(2, 6):
        c = ws3.cell(row=ri, column=ci, value='')
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center', vertical='center')
        if ci in [2]:
            c.number_format = '0.00'
        else:
            c.number_format = '0.00'
    ws3.row_dimensions[ri].height = 22

col_widths_3 = [18] + [14] * 4
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 4: 参数页
# ============================================================
ws4 = wb.create_sheet('参数页')
ws4.sheet_view.showGridLines = False

style_title(ws4, '系统参数页（批量生成文件时注入变量）', 1)
ws4.row_dimensions[1].height = 35

headers4 = ['参数名称', '参数值']
for i, h in enumerate(headers4, 1):
    c = ws4.cell(row=2, column=i, value=h)
    style_header(c)
ws4.row_dimensions[2].height = 30

params = [
    ['CURRENT_DATE', ''],
    ['CURRENT_YEAR', ''],
    ['CURRENT_MONTH', ''],
    ['CURRENT_ORG', ''],
    ['CURRENT_ORG_ID', ''],
    ['TEMPLATE_VERSION', 'V1.0'],
]
for ri, (k, v) in enumerate(params, 3):
    ck = ws4.cell(row=ri, column=1, value=k)
    ck.border = thin_border()
    ck.alignment = Alignment(horizontal='center', vertical='center')
    ck.font = Font(bold=True, color='2E4057')

    cv = ws4.cell(row=ri, column=2, value=v)
    cv.border = thin_border()
    cv.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[ri].height = 22

col_widths_4 = [25, 30]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# 说明行
ws4.merge_cells(start_row=10, start_column=1, end_row=10, end_column=2)
note = ws4.cell(row=10, column=1,
    value='注：CURRENT_DATE、CURRENT_YEAR、CURRENT_MONTH、CURRENT_ORG、CURRENT_ORG_ID '
         '在批量生成各机构月度文件时由系统自动注入，其他参数自定义。')
note.font = Font(italic=True, size=9, color='888888')
note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws4.row_dimensions[10].height = 30

# 保存
out_path = '/Users/zhaoyu/Desktop/机构绩效管理_考核体系模版.xlsx'
wb.save(out_path)
print(f'Excel template saved to: {out_path}')
