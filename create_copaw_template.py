#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def style_header(cell, bg='2E4057', fg='FFFFFF'):
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_title(ws, text, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=14, color='2E4057')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30

def thin_border():
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)

wb = Workbook()

# Sheet 1: ???
ws1 = wb.active
ws1.title = '???'
ws1.sheet_view.showGridLines = False

style_title(ws1, 'CoPaw???? - ????????', 1)

ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)
cell = ws1.cell(row=2, column=1, value='????????????????????Excel??????????????????')
cell.font = Font(italic=True, size=9, color='888888')
cell.alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[2].height = 20

headers = ['??', '??', '????', '????', '??', '??', '????', '????', '???', 
           '?????', '?????', '???????', '??????', '????', '????', '???']
for i, h in enumerate(headers, 1):
    c = ws1.cell(row=3, column=i, value=h)
    style_header(c)
ws1.row_dimensions[3].height = 35

sample_data = [
    ['????', '????', '????', '??????', 0.15, '??', 100.00, 8.33, 7.85, '=I4/G4', '=I4/H4', '=MAX(0,MIN(100,I4/H4*100))', '=L4*E4', '=SUM(M4:M5)', '=SUM(N4:N5)', '=SUM(O:O)'],
    ['????', '????', '????', '?????', 0.10, '%', 15.00, 1.25, 1.10, '=I5/G5', '=I5/H5', '=MAX(0,MIN(100,I5/H5*100))', '=L5*E5', '=SUM(M4:M5)', '=SUM(N4:N5)', '=SUM(O:O)'],
    ['????', '????', '?????', '???????', 0.20, '%', 1.50, 0.125, 0.11, '=I6/G6', '=I6/H6', '=MAX(0,MIN(100,(1-I6/G6)*100))', '=L6*E6', '=SUM(M6:M7)', '=SUM(N4:N7)', '=SUM(O:O)'],
    ['????', '????', '?????', '?????', 0.15, '%', 180.00, 15.00, 16.50, '=I7/G7', '=I7/H7', '=MAX(0,MIN(100,I7/G7*100))', '=L7*E7', '=SUM(M6:M7)', '=SUM(N4:N7)', '=SUM(O:O)'],
]

for ri, row_data in enumerate(sample_data, 4):
    for ci, val in enumerate(row_data, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.border = thin_border()
        if ci == 5:
            c.number_format = '0.00%'
        elif ci in [7, 8, 9]:
            c.number_format = '0.00'
        elif ci in [10, 11, 12, 13, 14, 15, 16]:
            c.number_format = '0.0000'
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[ri].height = 22

col_widths_1 = [12, 12, 16, 16, 8, 8, 12, 12, 10, 12, 12, 14, 12, 10, 10, 10]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Sheet 2: ???
ws2 = wb.create_sheet('???')
ws2.sheet_view.showGridLines = False

style_title(ws2, '???????', 1)
ws2.row_dimensions[1].height = 35

headers2 = ['????', '??ID', '????', '?????']
for i, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=i, value=h)
    style_header(c)
ws2.row_dimensions[2].height = 30

institutions = [
    ['????', 'ORG_BJ', '???', '???/EMP001'],
    ['????', 'ORG_SH', '???', '???/EMP002'],
    ['????', 'ORG_SZ', '???', '???/EMP003'],
    ['????', 'ORG_GZ', '???', '???/EMP004'],
]
for ri, row_data in enumerate(institutions, 3):
    for ci, val in enumerate(row_data, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[ri].height = 22

col_widths_2 = [20, 15, 15, 20]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Sheet 3: ?????
ws3 = wb.create_sheet('?????')
ws3.sheet_view.showGridLines = False

style_title(ws3, '????????????????', 1)
ws3.row_dimensions[1].height = 35

row2 = ['????', '??????', '?????', '?????', '?????']
row3 = ['', '??/EMP101', '??/EMP101', '???/EMP102', '???/EMP102']
row4 = ['', '??', '%', '%', '%']

for ci, val in enumerate(row2, 1):
    c = ws3.cell(row=2, column=ci, value=val)
    if ci == 1:
        style_header(c, '2E4057')
    else:
        style_header(c, '048A81')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for ci, val in enumerate(row3, 1):
    c = ws3.cell(row=3, column=ci, value=val)
    if ci == 1:
        style_header(c, '2E4057')
    else:
        style_header(c, '048A81')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for ci, val in enumerate(row4, 1):
    c = ws3.cell(row=4, column=ci, value=val)
    if ci == 1:
        style_header(c, '2E4057')
    else:
        style_header(c, '048A81')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws3.row_dimensions[2].height = 30
ws3.row_dimensions[3].height = 30
ws3.row_dimensions[4].height = 30

for ri, (org_name, org_id) in enumerate([
    ('????', 'ORG_BJ'), ('????', 'ORG_SH'), ('????', 'ORG_SZ'), ('????', 'ORG_GZ'),
], 5):
    c1 = ws3.cell(row=ri, column=1, value=org_name)
    style_header(c1, '2E4057')
    c1.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(2, 6):
        c = ws3.cell(row=ri, column=ci, value='')
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center', vertical='center')
        if ci in [2]:
            c.number_format = '0.00'
        else:
            c.number_format = '0.00'
    ws3.row_dimensions[ri].height = 22

col_widths_3 = [18] + [14] * 4
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Sheet 4: ???
ws4 = wb.create_sheet('???')
ws4.sheet_view.showGridLines = False

style_title(ws4, '??????????????????', 1)
ws4.row_dimensions[1].height = 35

headers4 = ['????', '???']
for i, h in enumerate(headers4, 1):
    c = ws4.cell(row=2, column=i, value=h)
    style_header(c)
ws4.row_dimensions[2].height = 30

params = [
    ['CURRENT_DATE', ''],
    ['CURRENT_YEAR', ''],
    ['CURRENT_MONTH', ''],
    ['CURRENT_ORG', ''],
    ['CURRENT_ORG_ID', ''],
    ['TEMPLATE_VERSION', 'V1.0'],
]
for ri, (k, v) in enumerate(params, 3):
    ck = ws4.cell(row=ri, column=1, value=k)
    ck.border = thin_border()
    ck.alignment = Alignment(horizontal='center', vertical='center')
    ck.font = Font(bold=True, color='2E4057')

    cv = ws4.cell(row=ri, column=2, value=v)
    cv.border = thin_border()
    cv.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[ri].height = 22

col_widths_4 = [25, 30]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells(start_row=10, start_column=1, end_row=10, end_column=2)
note = ws4.cell(row=10, column=1,
    value='??CURRENT_DATE?CURRENT_YEAR?CURRENT_MONTH?CURRENT_ORG?CURRENT_ORG_ID ?????????????????????????????')
note.font = Font(italic=True, size=9, color='888888')
note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws4.row_dimensions[10].height = 30

out_path = 'CoPaw????_??????.xlsx'
wb.save(out_path)
print(f'Excel??????: {out_path}')
